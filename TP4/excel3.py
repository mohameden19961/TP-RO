import openpyxl
from openpyxl import Workbook

wb = Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

ws_donnees = wb.create_sheet("Données")

# --- DONNÉES DE L'EXERCICE ---
m = 4  
n = 9  
longueur_bobine = 300  
largeurs = [80, 100, 140, 180]
demandes = [240, 190, 120, 100]
patrons = [
    [3, 0, 0, 1, 0, 2, 1, 2, 0],
    [0, 3, 0, 0, 1, 1, 2, 0, 1],
    [0, 0, 2, 0, 0, 0, 0, 1, 1],
    [0, 0, 0, 1, 1, 0, 0, 0, 0]
]
pertes = [60, 0, 20, 40, 20, 40, 20, 0, 60]

# --- REMPLISSAGE FEUILLE DONNÉES ---
ws_donnees['A1'] = "Nombre de types de rouleaux (m)"
ws_donnees['B1'] = m
ws_donnees['A2'] = "Nombre de patrons (n)"
ws_donnees['B2'] = n
ws_donnees['A3'] = "Longueur bobine (cm)"
ws_donnees['B3'] = longueur_bobine

ws_donnees['A5'], ws_donnees['B5'], ws_donnees['C5'] = "Type", "Largeur (cm)", "Demande"
for i in range(m):
    ws_donnees.cell(6+i, 1).value = i + 1
    ws_donnees.cell(6+i, 2).value = largeurs[i]
    ws_donnees.cell(6+i, 3).value = demandes[i]

ws_donnees['A11'] = "Matrice des patrons de découpe"
ws_donnees['A12'] = "Type/Patron"
for j in range(n):
    ws_donnees.cell(12, 2+j).value = f"P{j+1}"

for i in range(m):
    ws_donnees.cell(13+i, 1).value = f"{largeurs[i]} cm"
    for j in range(n):
        ws_donnees.cell(13+i, 2+j).value = patrons[i][j]

ws_donnees.cell(13+m, 1).value = "Perte (cm)"
for j in range(n):
    ws_donnees.cell(13+m, 2+j).value = pertes[j]

# --- REMPLISSAGE FEUILLE RÉSULTATS AVEC FORMULES ---
ws_resultats = wb.create_sheet("Résultats")

ws_resultats['A1'] = "Perte totale minimale (cm)"
# Formule : Somme de la colonne D (Perte totale par patron)
ws_resultats['B1'] = "=SUM(D5:D13)" 

ws_resultats['A2'] = "Nombre de bobines utilisées"
# Formule : Somme de la colonne B (Nombre de bobines par patron)
ws_resultats['B2'] = "=SUM(B5:B13)"

ws_resultats['A4'], ws_resultats['B4'], ws_resultats['C4'], ws_resultats['D4'] = "Patron", "Nb Bobines", "Perte Unit.", "Perte Totale"

for j in range(n):
    ws_resultats.cell(5+j, 1).value = f"P{j+1}"
    ws_resultats.cell(5+j, 2).value = 0  # Valeur par défaut à modifier manuellement
    ws_resultats.cell(5+j, 3).value = pertes[j]
    # Formule : Nb Bobines * Perte Unit.
    ws_resultats.cell(5+j, 4).value = f"=B{5+j}*C{5+j}"

# --- VÉRIFICATION DE LA DEMANDE ---
ws_resultats['A16'] = "Vérification de la demande"
ws_resultats['A17'], ws_resultats['B17'], ws_resultats['C17'], ws_resultats['D17'], ws_resultats['E17'] = "Type", "Largeur", "Demande", "Production", "Surplus"

for i in range(m):
    ws_resultats.cell(18+i, 1).value = i + 1
    ws_resultats.cell(18+i, 2).value = largeurs[i]
    ws_resultats.cell(18+i, 3).value = demandes[i]
    
    # Formule SOMMEPROD pour calculer la production réelle selon les patrons choisis
    # Exemple pour le premier type : SOMMEPROD(B5:B13; 'Données'!B13:J13)
    col_start = 'B'
    col_end = openpyxl.utils.get_column_letter(1 + n)
    ws_resultats.cell(18+i, 4).value = f"=SUMPRODUCT(B$5:B$13, 'Données'!{col_start}{13+i}:{col_end}{13+i})"
    
    # Formule : Production - Demande
    ws_resultats.cell(18+i, 5).value = f"=D{18+i}-C{18+i}"

# Ajustement colonnes
for sheet in [ws_donnees, ws_resultats]:
    for col in ['A', 'B', 'C', 'D', 'E']:
        sheet.column_dimensions[col].width = 18

filename = "exercice3.xlsx"
wb.save(filename)
print(f"Fichier '{filename}' prêt. Les formules de calcul sont intégrées !")