import openpyxl
from openpyxl import Workbook

wb = Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

ws_donnees = wb.create_sheet("Données")

n = 20  
capacity = 1000  

benefices = [54, 82, 69, 23, 29, 26, 86, 34, 18, 94, 99, 11, 99, 83, 99, 64, 10, 49, 40, 25]

poids = [66, 80, 77, 67, 83, 58, 52, 69, 97, 87, 90, 54, 71, 74, 78, 69, 89, 79, 75, 91]


ws_donnees['A1'] = "Nombre d'objets (n)"
ws_donnees['B1'] = n

ws_donnees['A2'] = "Capacité (kg)"
ws_donnees['B2'] = capacity

ws_donnees['A3'] = "Objet"
for i in range(n):
    ws_donnees.cell(3, 2+i).value = i + 1

ws_donnees['A4'] = "Bénéfice"
for i in range(n):
    ws_donnees.cell(4, 2+i).value = benefices[i]

ws_donnees['A5'] = "Poids (kg)"
for i in range(n):
    ws_donnees.cell(5, 2+i).value = poids[i]

ws_resultats = wb.create_sheet("Résultats")

ws_resultats['A1'] = "Bénéfice optimal"
ws_resultats['B1'] = ""  

ws_resultats['A2'] = "Objet"
for i in range(n):
    ws_resultats.cell(2, 2+i).value = i + 1

ws_resultats['A3'] = "Chargé (0/1)"
for i in range(n):
    ws_resultats.cell(3, 2+i).value = ""  

ws_donnees.column_dimensions['A'].width = 20
ws_resultats.column_dimensions['A'].width = 20

filename = "exercice1.xlsx" 
wb.save(filename)

print(f"Fichier '{filename}' créé avec succès!")
print(f"\nContenu de la page 'Données':")
print(f"  - Nombre d'objets: {n}")
print(f"  - Capacité: {capacity} kg")
print(f"  - Bénéfices: {benefices}")
print(f"  - Poids: {poids}")
print(f"\nLa page 'Résultats' a été préparée pour recevoir les solutions.")