import openpyxl
from openpyxl import Workbook

# Créer un nouveau classeur Excel
wb = Workbook()

# Supprimer la feuille par défaut
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Créer la page "Données"
ws_donnees = wb.create_sheet("Données")

# Données du problème (selon le cours RO_06, slide 3)
n = 20  # Nombre d'objets
capacity = 1000  # Capacité en kg (1 tonne = 1000 kg)

# Bénéfices des 20 objets
benefices = [54, 82, 69, 23, 29, 26, 86, 34, 18, 94, 99, 11, 99, 83, 99, 64, 10, 49, 40, 25]

# Poids des 20 objets en kg
poids = [66, 80, 77, 67, 83, 58, 52, 69, 97, 87, 90, 54, 71, 74, 78, 69, 89, 79, 75, 91]

# Remplir la page "Données"
# Ligne 1: Nombre d'objets
ws_donnees['A1'] = "Nombre d'objets (n)"
ws_donnees['B1'] = n

# Ligne 2: Capacité
ws_donnees['A2'] = "Capacité (kg)"
ws_donnees['B2'] = capacity

# Ligne 3: En-têtes pour les objets
ws_donnees['A3'] = "Objet"
for i in range(n):
    ws_donnees.cell(3, 2+i).value = i + 1

# Ligne 4: Bénéfices
ws_donnees['A4'] = "Bénéfice"
for i in range(n):
    ws_donnees.cell(4, 2+i).value = benefices[i]

# Ligne 5: Poids
ws_donnees['A5'] = "Poids (kg)"
for i in range(n):
    ws_donnees.cell(5, 2+i).value = poids[i]

# Créer la page "Résultats"
ws_resultats = wb.create_sheet("Résultats")

# En-têtes pour la page Résultats
ws_resultats['A1'] = "Bénéfice optimal"
ws_resultats['B1'] = ""  # Sera rempli par le programme de résolution

ws_resultats['A2'] = "Objet"
for i in range(n):
    ws_resultats.cell(2, 2+i).value = i + 1

ws_resultats['A3'] = "Chargé (0/1)"
for i in range(n):
    ws_resultats.cell(3, 2+i).value = ""  # Sera rempli par le programme de résolution

# Ajuster la largeur des colonnes pour une meilleure lisibilité
ws_donnees.column_dimensions['A'].width = 20
ws_resultats.column_dimensions['A'].width = 20

# Sauvegarder le fichier
filename = "exercice1.xlsx" 
wb.save(filename)

print(f"Fichier '{filename}' créé avec succès!")
print(f"\nContenu de la page 'Données':")
print(f"  - Nombre d'objets: {n}")
print(f"  - Capacité: {capacity} kg")
print(f"  - Bénéfices: {benefices}")
print(f"  - Poids: {poids}")
print(f"\nLa page 'Résultats' a été préparée pour recevoir les solutions.")