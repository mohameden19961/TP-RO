from ortools.linear_solver import pywraplp
import openpyxl

def TP3_Exercice2():
    """
    Résolution du problème du sac-à-dos borné (Bounded Knapsack Problem)
    pour le chargement d'un camion de capacité 10 tonnes.
    Bénéfice optimal attendu: 10 625
    """
    # Nom du fichier Excel
    file = "exercice2.xlsx"
    
    # Charger le classeur Excel
    wb = openpyxl.load_workbook(file)
    
    # Lire les données depuis la page "Données"
    ws = wb["Données"]
    n = ws.cell(1, 2).value  # Nombre de types d'objets (20)
    capacity = ws.cell(2, 2).value  # Capacité en kg (10000 = 10 tonnes)
    benefice = [ws.cell(4, 2+i).value for i in range(n)]  # Bénéfices
    poids = [ws.cell(5, 2+i).value for i in range(n)]  # Poids
    nombre = [ws.cell(6, 2+i).value for i in range(n)]  # Nombre d'objets disponibles
    
    print(f"Nombre de types d'objets: {n}")
    print(f"Capacité du camion: {capacity} kg ({capacity/1000} tonnes)")
    
    # Appel du solveur CBC pour les problèmes linéaires en nombres entiers
    solver = pywraplp.Solver('TP3_Exercice2', 
                             pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
    infinity = solver.infinity()
    
    # Création des variables de décision Xi (entières)
    # Xi = nombre d'objets de type i chargés (0 <= Xi <= mi)
    X = []
    for i in range(n):
        X.append(solver.IntVar(0, nombre[i], f'X[{i}]'))
    
    print(f'Nombre de variables = {solver.NumVariables()}')
    
    # Création de la fonction objectif: Maximiser Σ(bi * Xi)
    objective = solver.Objective()
    for i in range(n):
        objective.SetCoefficient(X[i], benefice[i])
    objective.SetMaximization()
    
    # Création de la contrainte de capacité: Σ(pi * Xi) <= Capacity
    contrainte_capacite = solver.Constraint(-infinity, capacity, 'capacite')
    for i in range(n):
        contrainte_capacite.SetCoefficient(X[i], poids[i])
    
    print(f'Nombre de contraintes = {solver.NumConstraints()}')
    
    # Résolution du problème
    status = solver.Solve()
    
    # Affichage des résultats
    if status == pywraplp.Solver.OPTIMAL:
        print('\n=== Solution Optimale Trouvée ===')
        print(f'Valeur optimale (bénéfice) = {solver.Objective().Value()}')
        
        # Calculer le poids total et le nombre total d'objets chargés
        poids_total = 0
        nombre_total_objets = 0
        
        print('\nObjets chargés:')
        for i in range(n):
            nb_charges = int(X[i].solution_value())
            if nb_charges > 0:
                poids_total += poids[i] * nb_charges
                nombre_total_objets += nb_charges
                print(f'  Type {i+1}: {nb_charges} objet(s) / {nombre[i]} disponible(s) - '
                      f'Bénéfice unitaire = {benefice[i]}, Poids unitaire = {poids[i]} kg')
        
        print(f'\nNombre total d\'objets chargés: {nombre_total_objets}')
        print(f'Poids total chargé: {poids_total} kg / {capacity} kg ({poids_total/1000:.2f} tonnes)')
        print(f'Capacité restante: {capacity - poids_total} kg ({(capacity - poids_total)/1000:.2f} tonnes)')
        
        # Écrire les résultats dans la page "Résultats"
        ws = wb["Résultats"]
        ws['B1'] = solver.Objective().Value()  # Bénéfice optimal
        
        # Écrire les valeurs des variables de décision
        for i in range(n):
            ws.cell(3, 2+i).value = int(X[i].solution_value())
        
        # Sauvegarder le classeur
        wb.save(file)
        print(f'\nRésultats sauvegardés dans {file}')
        
        # Informations avancées
        print('\n=== Informations avancées ===')
        print(f'Problème résolu en {solver.wall_time()} millisecondes')
        print(f'Problème résolu en {solver.iterations()} itérations')
        print(f'Problème résolu en {solver.nodes()} noeuds branch-and-bound')
        
    elif status == pywraplp.Solver.FEASIBLE:
        print('\nUne solution réalisable a été trouvée (mais pas nécessairement optimale)')
        print(f'Bénéfice = {solver.Objective().Value()}')
    else:
        print('\nAucune solution trouvée!')
    
    # Fermer le classeur
    wb.close()

# Exécution du programme
TP3_Exercice2()