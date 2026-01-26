from ortools.linear_solver import pywraplp
import openpyxl

# Nom du fichier Excel
file = "exercice1.xlsx"

# Charger le classeur Excel
wb = openpyxl.load_workbook(file)

# Lire les données depuis la page "Données"
ws = wb["Données"]
n = ws.cell(1, 2).value  # Nombre d'objets (20)
capacity = ws.cell(2, 2).value  # Capacité en kg (1000)
benefice = [ws.cell(4, 2+i).value for i in range(n)]  # Bénéfices
poids = [ws.cell(5, 2+i).value for i in range(n)]  # Poids

# Créer le solveur
solver = pywraplp.Solver.CreateSolver('SCIP')

if not solver:
    print("Erreur: Le solveur n'a pas pu être créé")
    exit()

# Variables de décision Xi (binaires)
# Xi = 1 si l'objet i est chargé, 0 sinon
X = []
for i in range(n):
    X.append(solver.BoolVar(f'X[{i}]'))

# Fonction objectif: Maximiser le bénéfice total
# Max Σ(bi * Xi)
objective = solver.Objective()
for i in range(n):
    objective.SetCoefficient(X[i], benefice[i])
objective.SetMaximization()

# Contrainte de capacité: Le poids total ne doit pas dépasser la capacité
# Σ(pi * Xi) <= Capacity
constraint_capacity = solver.Constraint(-solver.infinity(), capacity)
for i in range(n):
    constraint_capacity.SetCoefficient(X[i], poids[i])

# Résoudre le problème
status = solver.Solve()

# Vérifier le statut de la solution
if status == pywraplp.Solver.OPTIMAL:
    print("Solution optimale trouvée!")
    print(f"Bénéfice optimal: {solver.Objective().Value()}")
    
    # Afficher les objets sélectionnés
    poids_total = 0
    print("\nObjets chargés:")
    for i in range(n):
        if X[i].solution_value() == 1:
            print(f"  Objet {i+1}: Bénéfice = {benefice[i]}, Poids = {poids[i]} kg")
            poids_total += poids[i]
    print(f"\nPoids total chargé: {poids_total} kg sur {capacity} kg")
    
    # Écrire les résultats dans la page "Résultats"
    ws = wb["Résultats"]
    ws['B1'] = solver.Objective().Value()  # Bénéfice optimal
    
    # Écrire les valeurs des variables de décision
    for i in range(n):
        ws.cell(3, 2+i).value = X[i].solution_value()
    
    # Sauvegarder le classeur
    wb.save(file)
    print(f"\nRésultats sauvegardés dans {file}")
    
elif status == pywraplp.Solver.FEASIBLE:
    print("Une solution réalisable a été trouvée (mais peut-être pas optimale)")
    print(f"Bénéfice: {solver.Objective().Value()}")
else:
    print("Aucune solution trouvée")

# Fermer le classeur
wb.close()