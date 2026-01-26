import openpyxl
from openpyxl import Workbook

# Créer un nouveau classeur Excel
wb = Workbook()

# Supprimer la feuille par défaut
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Créer la page "Données"
ws_donnees = wb.create_sheet("Données")

# Données du problème (selon le cours RO_06, slide 5)
n = 20  # Nombre de types d'objets
capacity = 10000  # Capacité en kg (10 tonnes = 10000 kg)

# Bénéfices des 20 types d'objets
benefices = [54, 82, 69, 23, 29, 26, 86, 34, 18, 94, 99, 11, 99, 83, 99, 64, 10, 49, 40, 25]

# Poids des 20 types d'objets en kg
poids = [66, 80, 77, 67, 83, 58, 52, 69, 97, 87, 90, 54, 71, 74, 78, 69, 89, 79, 75, 91]

# Nombre d'objets disponibles de chaque type
nombre = [19, 9, 16, 12, 8, 18, 10, 10, 12, 10, 7, 13, 15, 14, 11, 20, 9, 13, 16, 18]

# Remplir la page "Données"
# Ligne 1: Nombre de types d'objets
ws_donnees['A1'] = "Nombre de types d'objets (n)"
ws_donnees['B1'] = n

# Ligne 2: Capacité
ws_donnees['A2'] = "Capacité (kg)"
ws_donnees['B2'] = capacity

# Ligne 3: En-têtes pour les types d'objets
ws_donnees['A3'] = "Type d'objet"
for i in range(n):
    ws_donnees.cell(3, 2+i).value = i + 1

# Ligne 4: Bénéfices
ws_donnees['A4'] = "Bénéfice"
for i in range(n):
    ws_donnees.cell(4, 2+i).value = benefices[i]

# Ligne 5: Poids
ws_donnees['A5'] = "Poids (kg)"
for i in range(n):
    ws_donnees.cell(5, 2+i).value = poids[i]

# Ligne 6: Nombre disponible
ws_donnees['A6'] = "Nombre disponible"
for i in range(n):
    ws_donnees.cell(6, 2+i).value = nombre[i]

# Créer la page "Résultats"
ws_resultats = wb.create_sheet("Résultats")

# En-têtes pour la page Résultats
ws_resultats['A1'] = "Bénéfice optimal"
ws_resultats['B1'] = ""  # Sera rempli par le programme de résolution

ws_resultats['A2'] = "Type d'objet"
for i in range(n):
    ws_resultats.cell(2, 2+i).value = i + 1

ws_resultats['A3'] = "Nombre chargé"
for i in range(n):
    ws_resultats.cell(3, 2+i).value = ""  # Sera rempli par le programme de résolution

# Ajuster la largeur des colonnes pour une meilleure lisibilité
ws_donnees.column_dimensions['A'].width = 30
ws_resultats.column_dimensions['A'].width = 20

# Sauvegarder le fichier
filename = "exercice2.xlsx"
wb.save(filename)

print(f"Fichier '{filename}' créé avec succès!")
print(f"\nContenu de la page 'Données':")
print(f"  - Nombre de types d'objets: {n}")
print(f"  - Capacité: {capacity} kg ({capacity/1000} tonnes)")
print(f"  - Bénéfices: {benefices}")
print(f"  - Poids: {poids}")
print(f"  - Nombre disponible: {nombre}")
print(f"\nLa page 'Résultats' a été préparée pour recevoir les solutions.")
print(f"\nBénéfice optimal attendu: 10 625")