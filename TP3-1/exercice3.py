from ortools.linear_solver import pywraplp
import openpyxl

def TP3_Exercice3():
    file = "exercice3.xlsx"
    wb = openpyxl.load_workbook(file)
    ws = wb["Données"]
    
    n = int(ws.cell(1, 2).value)
    m = int(ws.cell(2, 2).value)
    capacity = int(ws.cell(3, 2).value)
    
    benefice = [ws.cell(5, 2+i).value for i in range(n)]
    poids = [ws.cell(6, 2+i).value for i in range(n)]
    nombre = [ws.cell(7, 2+i).value for i in range(n)]
    
    print(f"Nombre de types d'objets: {n}")
    print(f"Nombre d'avions: {m}")
    print(f"Capacité par avion: {capacity} tonnes")
    
    solver = pywraplp.Solver('TP3_Exercice3', pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
    infinity = solver.infinity()
    
    X = [[solver.IntVar(0, nombre[i], f'X[{i}][{j}]') for j in range(m)] for i in range(n)]#:comment 
    
    objective = solver.Objective()
    for i in range(n):
        for j in range(m):
            objective.SetCoefficient(X[i][j], benefice[i])
    objective.SetMaximization()
    
    for j in range(m):
        contrainte_capacite = solver.Constraint(-infinity, capacity, f'capacite_{j}')
        for i in range(n):
            contrainte_capacite.SetCoefficient(X[i][j], poids[i])
    
    for i in range(n):
        contrainte_disponibilite = solver.Constraint(-infinity, nombre[i], f'disponibilite_{i}')
        for j in range(m):
            contrainte_disponibilite.SetCoefficient(X[i][j], 1)
    
    solver.Solve()
    
    print("Solution optimale trouvée!")
    print(f"Bénéfice optimal: {solver.Objective().Value()}")
    
    ws = wb["Résultats"]
    ws['B1'] = solver.Objective().Value()
    
    for i in range(n):
        for j in range(m):
            ws.cell(3+i, 2+j).value = int(X[i][j].solution_value())
    
    wb.save(file)
    print(f"Résultats sauvegardés dans {file}")
    wb.close()

TP3_Exercice3()