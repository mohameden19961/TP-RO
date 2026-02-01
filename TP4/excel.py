import openpyxl
from openpyxl import Workbook

wb = Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

ws_donnees = wb.create_sheet("Données")

# Données de l'exercice 1 - Bin Packing Flotte d'avions
n = 10  # nombre de types d'objets
nb_max_bins = 20  # nombre maximum de vols disponibles
capacity = 100  # capacité en tonnes

# Poids des objets en tonnes
poids = [5, 4, 3, 2, 9, 6, 8, 1, 4, 7]

# Nombre d'objets disponibles de chaque type
nombre = [15, 52, 39, 18, 45, 27, 12, 12, 14, 31]

# Page Données
ws_donnees['A1'] = "Nombre de types d'objets (n)"
ws_donnees['B1'] = n

ws_donnees['A2'] = "Capacité par avion (tonnes)"
ws_donnees['B2'] = capacity

ws_donnees['A3'] = "Nombre max de vols"
ws_donnees['B3'] = nb_max_bins

ws_donnees['A5'] = "Type d'objet"
for i in range(n):
    ws_donnees.cell(5, 2+i).value = i + 1

ws_donnees['A6'] = "Poids (tonnes)"
for i in range(n):
    ws_donnees.cell(6, 2+i).value = poids[i]

ws_donnees['A7'] = "Nombre disponible"
for i in range(n):
    ws_donnees.cell(7, 2+i).value = nombre[i]

# Page Résultats
ws_resultats = wb.create_sheet("Résultats")

ws_resultats['A1'] = "Nombre minimum de vols"
ws_resultats['B1'] = ""  # Sera rempli par le solveur

ws_resultats['A3'] = "Vol"
ws_resultats['B3'] = "Charge totale (tonnes)"
for i in range(n):
    ws_resultats.cell(3, 3+i).value = f"Type {i + 1}"

# Préparer les lignes pour les résultats de chaque vol
for j in range(nb_max_bins):
    ws_resultats.cell(4+j, 1).value = j + 1
    ws_resultats.cell(4+j, 2).value = ""  # Charge totale
    for i in range(n):
        ws_resultats.cell(4+j, 3+i).value = ""  # Nombre d'objets de type i

# Largeur des colonnes
ws_donnees.column_dimensions['A'].width = 25
ws_donnees.column_dimensions['B'].width = 15
ws_resultats.column_dimensions['A'].width = 15
ws_resultats.column_dimensions['B'].width = 25

filename = "exercice1.xlsx"
wb.save(filename)

print(f"Fichier '{filename}' créé avec succès!")
print(f"\nContenu de la page 'Données':")
print(f"  - Nombre de types d'objets: {n}")
print(f"  - Capacité par avion: {capacity} tonnes")
print(f"  - Nombre max de vols: {nb_max_bins}")
print(f"  - Poids: {poids}")
print(f"  - Nombre d'objets disponibles: {nombre}")
print(f"\nLa page 'Résultats' a été préparée pour recevoir les solutions.")
print(f"Solution optimale attendue: 14 vols")