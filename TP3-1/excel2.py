import openpyxl
from openpyxl import Workbook

wb = Workbook()

if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

ws_donnees = wb.create_sheet("Données")

n = 20  
capacity = 10000 

benefices = [54, 82, 69, 23, 29, 26, 86, 34, 18, 94, 99, 11, 99, 83, 99, 64, 10, 49, 40, 25]

poids = [66, 80, 77, 67, 83, 58, 52, 69, 97, 87, 90, 54, 71, 74, 78, 69, 89, 79, 75, 91]

nombre = [19, 9, 16, 12, 8, 18, 10, 10, 12, 10, 7, 13, 15, 14, 11, 20, 9, 13, 16, 18]


ws_donnees['A1'] = "Nombre de types d'objets (n)"
ws_donnees['B1'] = n

ws_donnees['A2'] = "Capacité (kg)"
ws_donnees['B2'] = capacity

ws_donnees['A3'] = "Type d'objet"
for i in range(n):
    ws_donnees.cell(3, 2+i).value = i + 1

ws_donnees['A4'] = "Bénéfice"
for i in range(n):
    ws_donnees.cell(4, 2+i).value = benefices[i]

ws_donnees['A5'] = "Poids (kg)"
for i in range(n):
    ws_donnees.cell(5, 2+i).value = poids[i]

ws_donnees['A6'] = "Nombre disponible"
for i in range(n):
    ws_donnees.cell(6, 2+i).value = nombre[i]

ws_resultats = wb.create_sheet("Résultats")

ws_resultats['A1'] = "Bénéfice optimal"
ws_resultats['B1'] = ""  

ws_resultats['A2'] = "Type d'objet"
for i in range(n):
    ws_resultats.cell(2, 2+i).value = i + 1

ws_resultats['A3'] = "Nombre chargé"
for i in range(n):
    ws_resultats.cell(3, 2+i).value = ""  

ws_donnees.column_dimensions['A'].width = 30
ws_resultats.column_dimensions['A'].width = 20

filename = "exercice2.xlsx"
wb.save(filename)

print(f"Fichier '{filename}' créé avec succès!")
print(f"\nContenu de la page 'Données':")
print(f"  - Nombre de types d'objets: {n}")
print(f"  - Capacité: {capacity} kg ({capacity/1000} tonnes)")
print(f"  - Bénéfices: {benefices}")
print(f"  - Poids: {poids}")
print(f"  - Nombre disponible: {nombre}")
print(f"\nLa page 'Résultats' a été préparée pour recevoir les solutions.")
print(f"\nBénéfice optimal attendu: 10 625")