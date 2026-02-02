from ortools.linear_solver import pywraplp
import openpyxl

def TP4_Exercice2():
    file = "exercice2.xlsx"
    wb = openpyxl.load_workbook(file)
    ws = wb["Données"]
    
    m = ws.cell(1, 2).value  
    n = ws.cell(2, 2).value  
    longueur_bobine = ws.cell(3, 2).value  
    
    largeurs = [ws.cell(6+i, 2).value for i in range(m)]
    demandes = [ws.cell(6+i, 3).value for i in range(m)]
    
    A = []
    for i in range(m):
        row = [ws.cell(13+i, 2+j).value for j in range(n)]
        A.append(row)
    
    pertes = [ws.cell(13+m, 2+j).value for j in range(n)]
    
    print(f"Nombre de types de rouleaux: {m}")
    print(f"Nombre de patrons: {n}")
    print(f"Longueur bobine: {longueur_bobine} cm")
    
    solver = pywraplp.Solver('TP4_Exercice2', pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
    infinity = solver.infinity()
    
    x = {}
    for j in range(n):
        x[j] = solver.IntVar(0, infinity, f'x[{j}]')
    
    print(f"Nombre de variables: {solver.NumVariables()}")
    
    objective = solver.Objective()
    for j in range(n):
        objective.SetCoefficient(x[j], 1)
    objective.SetMinimization()
    
    for i in range(m):
        contrainte_demande = solver.Constraint(demandes[i], infinity, f'demande_{i}')
        for j in range(n):
            contrainte_demande.SetCoefficient(x[j], A[i][j])
    
    print(f"Nombre de contraintes: {solver.NumConstraints()}")
    
    print("\nRésolution en cours...")
    status = solver.Solve()
    
    if status == pywraplp.Solver.OPTIMAL:
        print("\nSolution optimale trouvée!")
        nb_bobines = int(solver.Objective().Value())
        print(f"Nombre minimum de bobines: {nb_bobines}")
        print(f"Temps de calcul: {solver.wall_time()} ms")
        print(f"Itérations: {solver.iterations()}")
        
        perte_totale = sum(x[j].solution_value() * pertes[j] for j in range(n))
        print(f"Perte totale: {int(perte_totale)} cm")
        
        ws_res = wb["Résultats"]
        ws_res['B1'] = nb_bobines
        ws_res['B2'] = int(perte_totale)
        
        for j in range(n):
            nb_bobines_patron = int(x[j].solution_value())
            ws_res.cell(5+j, 2).value = nb_bobines_patron
            ws_res.cell(5+j, 4).value = nb_bobines_patron * pertes[j]
        
        for i in range(m):
            production = sum(A[i][j] * x[j].solution_value() for j in range(n))
            ws_res.cell(18+i, 4).value = int(production)
            ws_res.cell(18+i, 5).value = int(production - demandes[i])
        
        wb.save(file)
        print(f"\nRésultats sauvegardés dans {file}")
        
    elif status == pywraplp.Solver.FEASIBLE:
        print("Une solution réalisable a été trouvée (mais peut-être pas optimale)")
        print(f"Nombre de bobines: {int(solver.Objective().Value())}")
    else:
        print("Aucune solution trouvée")
        if status == pywraplp.Solver.INFEASIBLE:
            print("Le problème est infaisable.")
        elif status == pywraplp.Solver.UNBOUNDED:
            print("Le problème est non borné.")
    
    wb.close()

TP4_Exercice2()