import openpyxl
from openpyxl import Workbook

wb = Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

ws_donnees = wb.create_sheet("Données")

# Données de l'exercice 2 - Problème de découpe des bobines
m = 4  # nombre de types de rouleaux
n = 9  # nombre de patrons de découpe
longueur_bobine = 300  # en cm

# Largeurs des rouleaux demandés (en cm)
largeurs = [80, 100, 140, 180]

# Demandes pour chaque type de rouleau
demandes = [240, 190, 120, 100]

# Matrice des patrons de découpe (lignes = types de rouleaux, colonnes = patrons)
patrons = [
    [3, 0, 0, 1, 0, 2, 1, 2, 0],  # 80 cm
    [0, 3, 0, 0, 1, 1, 2, 0, 1],  # 100 cm
    [0, 0, 2, 0, 0, 0, 0, 1, 1],  # 140 cm
    [0, 0, 0, 1, 1, 0, 0, 0, 0]   # 180 cm
]

# Perte pour chaque patron (en cm)
pertes = [60, 0, 20, 40, 20, 40, 20, 0, 60]

# Page Données
ws_donnees['A1'] = "Nombre de types de rouleaux (m)"
ws_donnees['B1'] = m

ws_donnees['A2'] = "Nombre de patrons (n)"
ws_donnees['B2'] = n

ws_donnees['A3'] = "Longueur bobine (cm)"
ws_donnees['B3'] = longueur_bobine

# Tableau de la demande
ws_donnees['A5'] = "Type de rouleau"
ws_donnees['B5'] = "Largeur (cm)"
ws_donnees['C5'] = "Demande"

for i in range(m):
    ws_donnees.cell(6+i, 1).value = i + 1
    ws_donnees.cell(6+i, 2).value = largeurs[i]
    ws_donnees.cell(6+i, 3).value = demandes[i]

# Matrice des patrons de découpe
ws_donnees['A11'] = "Matrice des patrons de découpe"

ws_donnees['A12'] = "Type/Patron"
for j in range(n):
    ws_donnees.cell(12, 2+j).value = f"P{j+1}"

for i in range(m):
    ws_donnees.cell(13+i, 1).value = f"{largeurs[i]} cm"
    for j in range(n):
        ws_donnees.cell(13+i, 2+j).value = patrons[i][j]

# Ligne des pertes
ws_donnees.cell(13+m, 1).value = "Perte (cm)"
for j in range(n):
    ws_donnees.cell(13+m, 2+j).value = pertes[j]

# Page Résultats
ws_resultats = wb.create_sheet("Résultats")

ws_resultats['A1'] = "Nombre minimum de bobines"
ws_resultats['B1'] = ""  # Sera rempli par le solveur

ws_resultats['A2'] = "Perte totale (cm)"
ws_resultats['B2'] = ""  # Sera rempli par le solveur

ws_resultats['A4'] = "Patron"
ws_resultats['B4'] = "Nombre de bobines"
ws_resultats['C4'] = "Perte unitaire (cm)"
ws_resultats['D4'] = "Perte totale (cm)"

for j in range(n):
    ws_resultats.cell(5+j, 1).value = f"P{j+1}"
    ws_resultats.cell(5+j, 2).value = ""  # Nombre de bobines
    ws_resultats.cell(5+j, 3).value = pertes[j]
    ws_resultats.cell(5+j, 4).value = ""  # Perte totale

# Vérification de la demande
ws_resultats['A16'] = "Vérification de la demande"
ws_resultats['A17'] = "Type"
ws_resultats['B17'] = "Largeur (cm)"
ws_resultats['C17'] = "Demande"
ws_resultats['D17'] = "Production"
ws_resultats['E17'] = "Surplus"

for i in range(m):
    ws_resultats.cell(18+i, 1).value = i + 1
    ws_resultats.cell(18+i, 2).value = largeurs[i]
    ws_resultats.cell(18+i, 3).value = demandes[i]
    ws_resultats.cell(18+i, 4).value = ""  # Production
    ws_resultats.cell(18+i, 5).value = ""  # Surplus

# Largeur des colonnes
ws_donnees.column_dimensions['A'].width = 25
ws_donnees.column_dimensions['B'].width = 15
ws_donnees.column_dimensions['C'].width = 15
ws_resultats.column_dimensions['A'].width = 25
ws_resultats.column_dimensions['B'].width = 20
ws_resultats.column_dimensions['C'].width = 20
ws_resultats.column_dimensions['D'].width = 20
ws_resultats.column_dimensions['E'].width = 15

filename = "exercice2.xlsx"
wb.save(filename)

print(f"Fichier '{filename}' créé avec succès!")
print(f"\nContenu de la page 'Données':")
print(f"  - Nombre de types de rouleaux: {m}")
print(f"  - Nombre de patrons: {n}")
print(f"  - Longueur bobine: {longueur_bobine} cm")
print(f"  - Largeurs: {largeurs}")
print(f"  - Demandes: {demandes}")
print(f"  - Pertes par patron: {pertes}")
print(f"\nLa page 'Résultats' a été préparée pour recevoir les solutions.")
print(f"Solution optimale attendue: 250 bobines")