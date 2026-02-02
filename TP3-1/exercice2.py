from ortools.linear_solver import pywraplp
import openpyxl

def TP3_Exercice2():
    file = "exercice2.xlsx"
    wb = openpyxl.load_workbook(file)
    ws = wb["Données"]
    n = ws.cell(1, 2).value 
    capacity = ws.cell(2, 2).value  
    benefice = [ws.cell(4, 2+i).value for i in range(n)] 
    poids = [ws.cell(5, 2+i).value for i in range(n)]  
    nombre = [ws.cell(6, 2+i).value for i in range(n)]
    
    print(f"Nombre de types d'objets: {n}")
    print(f"Capacité du camion: {capacity} kg ({capacity/1000} tonnes)")
    
    solver = pywraplp.Solver('TP3_Exercice2',pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
    infinity = solver.infinity()

    X = []
    for i in range(n):
        X.append(solver.IntVar(0, nombre[i], f'X[{i}]'))
    
    
    objective = solver.Objective()
    for i in range(n):
        objective.SetCoefficient(X[i], benefice[i])
    objective.SetMaximization()
    
    contrainte_capacite = solver.Constraint(-infinity, capacity, 'capacite')
    for i in range(n):
        contrainte_capacite.SetCoefficient(X[i], poids[i])
    
    
    
    status = solver.Solve()
    
    if status == pywraplp.Solver.OPTIMAL:
        print('\n=== Solution Optimale Trouvée ===')
        print(f'Valeur optimale (bénéfice) = {solver.Objective().Value()}')
        
        poids_total = 0
        nombre_total_objets = 0
        
        
        ws = wb["Résultats"]
        ws['B1'] = solver.Objective().Value()  
        
        for i in range(n):
            ws.cell(3, 2+i).value = int(X[i].solution_value())
        
        wb.save(file)
        print(f'\nRésultats sauvegardés dans {file}')
    
    elif status == pywraplp.Solver.FEASIBLE:
        print('\nUne solution réalisable a été trouvée (mais pas nécessairement optimale)')
        print(f'Bénéfice = {solver.Objective().Value()}')
    else:
        print('\nAucune solution trouvée!')
    
    wb.close()

TP3_Exercice2()