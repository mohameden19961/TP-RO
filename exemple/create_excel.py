import openpyxl
from openpyxl import Workbook

wb = Workbook()

if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

ws_donnees = wb.create_sheet("Données")

n = 10
m = 10
capacity = 100

benefices = [60, 55, 42, 36, 98, 70, 82, 29, 80, 78]
poids = [5, 4, 3, 2, 9, 6, 8, 1, 4, 7]
nombre = [15, 52, 39, 18, 45, 27, 12, 12, 14, 31]

ws_donnees['A1'] = "Nombre de types d'objets (n)"
ws_donnees['B1'] = n
ws_donnees['A2'] = "Nombre d'avions (m)"
ws_donnees['B2'] = m
ws_donnees['A3'] = "Capacité par avion (tonnes)"
ws_donnees['B3'] = capacity

ws_donnees['A4'] = "Type d'objet"
for i in range(n):
    ws_donnees.cell(4, 2+i).value = i + 1

ws_donnees['A5'] = "Bénéfice (millions MRU)"
for i in range(n):
    ws_donnees.cell(5, 2+i).value = benefices[i]

ws_donnees['A6'] = "Poids (tonnes)"
for i in range(n):
    ws_donnees.cell(6, 2+i).value = poids[i]

ws_donnees['A7'] = "Nombre disponible"
for i in range(n):
    ws_donnees.cell(7, 2+i).value = nombre[i]

ws_donnees.column_dimensions['A'].width = 30

ws_resultats = wb.create_sheet("Résultats")

ws_resultats['A1'] = "Bénéfice optimal"
ws_resultats['B1'] = ""

ws_resultats['A2'] = "Avion"
for j in range(m):
    ws_resultats.cell(2, 2+j).value = j + 1

for i in range(n):
    ws_resultats.cell(3+i, 1).value = f"Type {i+1}"

ws_resultats.column_dimensions['A'].width = 15

filename = "exercice3.xlsx"
wb.save(filename)
print(f"Fichier '{filename}' créé avec succès!")