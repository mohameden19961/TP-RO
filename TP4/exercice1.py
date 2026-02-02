from ortools.linear_solver import pywraplp
import openpyxl

def TP4_Exercice1():
    file = "exercice1.xlsx"
    wb = openpyxl.load_workbook(file)
    ws = wb["Données"]
    
    n = ws.cell(1, 2).value  
    capacity = ws.cell(2, 2).value  
    nb_max_bins = ws.cell(3, 2).value  
    
    poids = [ws.cell(6, 2+i).value for i in range(n)]  
    nombre = [ws.cell(7, 2+i).value for i in range(n)]  
    
    print(f"Nombre de types d'objets: {n}")
    print(f"Capacité par avion: {capacity} tonnes")
    print(f"Nombre max de vols: {nb_max_bins}")
    
    solver = pywraplp.Solver('TP4_Exercice1', pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
    infinity = solver.infinity()
    
    
    x = {}
    for i in range(n):
        for j in range(nb_max_bins):
            x[i, j] = solver.IntVar(0, nombre[i], f'x[{i},{j}]')
    
    y = {}
    for j in range(nb_max_bins):
        y[j] = solver.IntVar(0, 1, f'y[{j}]')
    
    print(f"Nombre de variables: {solver.NumVariables()}")
    
    objective = solver.Objective()
    for j in range(nb_max_bins):
        objective.SetCoefficient(y[j], 1)
    objective.SetMinimization()
    
    
    for i in range(n):
        contrainte_demande = solver.Constraint(nombre[i], nombre[i], f'demande_{i}')
        for j in range(nb_max_bins):
            contrainte_demande.SetCoefficient(x[i, j], 1)
    
    for j in range(nb_max_bins):
        contrainte_capacite = solver.Constraint(-infinity, capacity, f'capacite_{j}')
        for i in range(n):
            contrainte_capacite.SetCoefficient(x[i, j], poids[i])
    
    for i in range(n):
        for j in range(nb_max_bins):
            contrainte_lien = solver.Constraint(-infinity, 0, f'lien_{i}_{j}')
            contrainte_lien.SetCoefficient(x[i, j], 1)
            contrainte_lien.SetCoefficient(y[j], -nombre[i])
    
    print(f"Nombre de contraintes: {solver.NumConstraints()}")
    
    status = solver.Solve()
    
    if status == pywraplp.Solver.OPTIMAL:
        print("\nSolution optimale trouvée!")
        print(f"Nombre minimum de vols: {int(solver.Objective().Value())}")
        print(f"Temps de calcul: {solver.wall_time()} ms")
        print(f"Itérations: {solver.iterations()}")
        
        ws_res = wb["Résultats"]
        ws_res['B1'] = int(solver.Objective().Value())
        
        row = 4
        vols_utilises = 0
        
        for j in range(nb_max_bins):
            if y[j].solution_value() > 0.5:  
                vols_utilises += 1
                
                ws_res.cell(row, 1).value = vols_utilises
                
                charge_totale = sum(x[i, j].solution_value() * poids[i] for i in range(n))
                ws_res.cell(row, 2).value = charge_totale
                
                for i in range(n):
                    nb_objets = int(x[i, j].solution_value())
                    ws_res.cell(row, 3+i).value = nb_objets
                
                row += 1
        
        wb.save(file)
        print(f"\nRésultats sauvegardés dans {file}")
        
    elif status == pywraplp.Solver.FEASIBLE:
        print("Une solution réalisable a été trouvée (mais peut-être pas optimale)")
        print(f"Nombre de vols: {int(solver.Objective().Value())}")
    else:
        print("Aucune solution trouvée")
        if status == pywraplp.Solver.INFEASIBLE:
            print("Le problème est infaisable.")
        elif status == pywraplp.Solver.UNBOUNDED:
            print("Le problème est non borné.")
    
    wb.close()

TP4_Exercice1()